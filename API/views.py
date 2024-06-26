from openpyxl.utils import get_column_letter
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import Student
from .serializers import StudentSerializer
from django.db import connections
from django.http import HttpResponse
import openpyxl


class GetStudents(APIView):
    def get(self, request):
        students = Student.objects.all()
        serializer = StudentSerializer(students, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class GetStudentDetails(APIView):
    def get(self, request, student_id):
        try:
            student = Student.objects.get(pk=student_id)
        except Student.DoesNotExist:
            return Response({"error": "Student not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = StudentSerializer(student)
        return Response(serializer.data, status=status.HTTP_200_OK)


class UpdateStudentStatus(APIView):
    def patch(self, request, data):
        try:
            new_data = data.split('&')
            student = Student.objects.get(pk=int(new_data[0]))
            student.status = new_data[1]
            student.save()
            return Response({"message": "Status updated"}, status=status.HTTP_200_OK)
        except Student.DoesNotExist:
            return Response({"message": "Student not found"}, status=status.HTTP_404_NOT_FOUND)


class ExportStudentsXLSX(APIView):
    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Students'

        headers = ['ФИО', 'Статус принятия', 'Номер телефона', 'VK', 'Email', 'Учебное заведение', 'Направление', 'Курс', 'Академ. степень', 'Отправил ли тесты', 'Telegram ID']
        ws.append(headers)

        column_widths = [35, 20, 15, 30, 30, 20, 30, 10, 20, 20, 15]

        for i, column_width in enumerate(column_widths, start=1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = column_width

        for student in Student.objects.all():
            row = [
                student.full_name,
                student.status,
                student.phone,
                student.vk,
                student.email,
                student.university,
                student.speciality,
                student.course,
                student.degree,
                'Да' if student.is_test_send else 'Нет',
                student.telegram
            ]
            ws.append(row)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=students.xlsx'
        wb.save(response)
        return response


def check_database_connection():
    try:
        with connections['default'].cursor() as cursor:
            cursor.execute("SELECT 1")
            result = cursor.fetchone()
            if result:
                print("Подключение к базе данных успешно!")
    except Exception as e:
        print(f"Ошибка подключения к базе данных: {str(e)}")
